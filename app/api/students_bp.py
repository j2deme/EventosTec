from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required
import requests
from app import db
from app.schemas import student_schema, students_schema
from app.models.student import Student
from app.utils.auth_helpers import require_admin
from openpyxl import Workbook
from typing import Any
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime, timezone
from app.utils.datetime_utils import localize_naive_datetime, safe_iso
from flask import current_app


# use centralized safe_iso from app.utils.datetime_utils


students_bp = Blueprint('students', __name__, url_prefix='/api/students')


@students_bp.route('/search', methods=['GET'])
def search_students():
    q = request.args.get('q', '').strip()
    if not q or len(q) < 2:
        return jsonify([])
    # Realizar búsqueda por nombre cuando la consulta es suficiente
    results = Student.query.filter(
        Student.full_name.ilike(f'%{q}%')).limit(10).all()
    return jsonify([
        {'id': s.id, 'full_name': s.full_name, 'control_number': s.control_number}
        for s in results
    ])


@students_bp.route('/', methods=['GET'])
def get_students():
    try:
        # Parámetros de búsqueda y paginación
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search = request.args.get('search', '')

        # Nuevos filtros
        event_id = request.args.get('event_id', type=int)
        activity_id = request.args.get('activity_id', type=int)
        career = request.args.get('career', '')

        query = Student.query

        # Filtro por evento o actividad (require joins)
        if event_id or activity_id:
            from app.models.registration import Registration
            from app.models.activity import Activity

            query = query.join(
                Registration, Registration.student_id == Student.id)
            query = query.join(Activity, Activity.id ==
                               Registration.activity_id)

            if event_id:
                query = query.filter(Activity.event_id == event_id)

            if activity_id:
                query = query.filter(Activity.id == activity_id)

            # Eliminar duplicados cuando hay joins
            query = query.distinct()

        # Filtro por carrera
        if career:
            query = query.filter(Student.career.ilike(f"%{career}%"))

        # Búsqueda general
        if search:
            search_filter = f"%{search}%"
            query = query.filter(
                db.or_(
                    Student.control_number.ilike(search_filter),
                    Student.full_name.ilike(search_filter),
                    Student.career.ilike(search_filter)
                )
            )

        # Ordenar por nombre
        query = query.order_by(Student.full_name)

        students = query.paginate(
            page=page, per_page=per_page, error_out=False
        )

        return jsonify({
            'students': students_schema.dump(students.items),
            'total': students.total,
            'pages': students.pages,
            'current_page': page
        }), 200

    except Exception as e:
        return jsonify({'message': 'Error al obtener estudiantes', 'error': str(e)}), 500

# Obtener estudiante por ID


@students_bp.route('/<int:student_id>', methods=['GET'])
def get_student(student_id):
    try:
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'message': 'Estudiante no encontrado'}), 404

        return jsonify({'student': student_schema.dump(student)}), 200

    except Exception as e:
        return jsonify({'message': 'Error al obtener estudiante', 'error': str(e)}), 500

# Buscar estudiante en sistema externo


@students_bp.route('/external-search', methods=['GET'])
@jwt_required()
def search_external_student():
    try:
        control_number = request.args.get('control_number')
        if not control_number:
            return jsonify({'message': 'Número de control es requerido'}), 400

        # Consultar sistema externo
        external_api_url = f"http://apps.tecvalles.mx:8091/api/estudiantes?search={control_number}"

        try:
            response = requests.get(external_api_url, timeout=10)
            if response.status_code == 200:
                return jsonify({'student': response.json()}), 200
            else:
                return jsonify({'message': 'Estudiante no encontrado en sistema externo'}), 404
        except requests.exceptions.RequestException:
            return jsonify({'message': 'Error de conexión con sistema externo'}), 503

    except Exception as e:
        return jsonify({'message': 'Error en búsqueda externa', 'error': str(e)}), 500

# Importar estudiante desde sistema externo


@students_bp.route('/import-external/<control_number>', methods=['POST'])
@jwt_required()
@require_admin
def import_external_student(control_number):
    try:
        # Consultar sistema externo
        external_api_url = f"http://apps.tecvalles.mx:8091/api/estudiantes?search={control_number}"

        try:
            response = requests.get(external_api_url, timeout=10)
            if response.status_code == 200:
                external_data = response.json()

                if external_data and len(external_data) > 0:
                    student_info = external_data[0]

                    # Verificar si ya existe
                    student = Student.query.filter_by(
                        control_number=control_number).first()
                    if not student:
                        # Crear nuevo estudiante
                        student = Student()
                        student.control_number = control_number
                        student.full_name = student_info.get('nombre', '')
                        student.career = student_info.get('carrera', '')
                        student.email = student_info.get('email', '')
                        db.session.add(student)
                        db.session.commit()

                        return jsonify({
                            'message': 'Estudiante importado exitosamente',
                            'student': student_schema.dump(student)
                        }), 201
                    else:
                        return jsonify({
                            'message': 'Estudiante ya existe en el sistema',
                            'student': student_schema.dump(student)
                        }), 200
                else:
                    return jsonify({'message': 'Estudiante no encontrado en sistema externo'}), 404
            else:
                return jsonify({'message': 'Error al consultar sistema externo'}), 503
        except requests.exceptions.RequestException:
            return jsonify({'message': 'Error de conexión con sistema externo'}), 503

    except Exception as e:
        db.session.rollback()
        return jsonify({'message': 'Error al importar estudiante', 'error': str(e)}), 500


# Endpoint proxy para validación externa usada por el modal de Walk-in (no requiere auth)
@students_bp.route('/validate', methods=['GET'])
def validate_student_proxy():
    """Proxy público que consulta el servicio externo de validación de estudiantes.

    Query params:
      - control_number (or username)

    Returns standardized JSON: { student: { control_number, full_name, career, email } }
    or 404 if not found, 503 on external errors.
    """
    control = request.args.get(
        'control_number') or request.args.get('username')
    if not control:
        return jsonify({'message': 'control_number es requerido'}), 400

    external_api = f"http://apps.tecvalles.mx:8091/api/validate/student?username={control}"
    try:
        resp = requests.get(external_api, timeout=8)
    except requests.exceptions.RequestException:
        return jsonify({'message': 'Error conectando al servicio externo'}), 503

    if resp.status_code == 200:
        try:
            data = resp.json()
        except Exception:
            return jsonify({'message': 'Respuesta externa inválida'}), 502

        if not isinstance(data, dict):
            return jsonify({'message': 'Respuesta externa inválida'}), 502

        # Some external services wrap payload in { success: true, data: { ... } }
        if isinstance(data, dict) and 'data' in data and isinstance(data.get('data'), dict):
            data = data.get('data')

        # Work with a local dict reference to satisfy static analysis
        d = data if isinstance(data, dict) else {}

        # Normalize keys if possible; career may be an object
        career = d.get('career') or d.get('carrera') or {}
        career_name = None
        if isinstance(career, dict):
            career_name = career.get('name') or career.get('nombre') or None
        else:
            career_name = career

        student = {
            'control_number': d.get('username') or d.get('control_number') or control,
            'full_name': d.get('name') or d.get('full_name') or d.get('nombre'),
            'career': career_name,
            'email': d.get('email') or ''
        }
        return jsonify({'student': student}), 200
    elif resp.status_code == 404:
        return jsonify({'message': 'Estudiante no encontrado'}), 404
    else:
        return jsonify({'message': 'Error desde servicio externo'}), 503

# Obtener actividades de un estudiante


@students_bp.route('/<int:student_id>/activities', methods=['GET'])
def get_student_activities(student_id):
    try:
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'message': 'Estudiante no encontrado'}), 404

        # Obtener actividades a través de asistencias y preregistros
        from app.models.attendance import Attendance
        from app.models.registration import Registration
        from app.models.activity import Activity

        # Actividades con asistencia
        attendance_activities = Activity.query.join(Attendance).filter(
            Attendance.student_id == student_id
        ).all()

        # Actividades con preregistro
        registration_activities = Activity.query.join(Registration).filter(
            Registration.student_id == student_id
        ).all()

        # Combinar y eliminar duplicados
        all_activities = list(
            set(attendance_activities + registration_activities))

        from app.schemas import activities_schema
        return jsonify({
            'activities': activities_schema.dump(all_activities)
        }), 200

    except Exception as e:
        return jsonify({'message': 'Error al obtener actividades del estudiante', 'error': str(e)}), 500


# Obtener horas acumuladas por evento de un estudiante
@students_bp.route('/<int:student_id>/hours-by-event', methods=['GET'])
def get_student_hours_by_event(student_id):
    """
    Calcula las horas confirmadas de un estudiante agrupadas por evento.
    Solo cuenta registros con status='Asistió'.
    """
    try:
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'message': 'Estudiante no encontrado'}), 404

        from app.models.registration import Registration
        from app.models.activity import Activity
        from app.models.event import Event
        from sqlalchemy import func

        # Query: agrupar por evento, sumar horas de actividades donde status='Asistió'
        results = db.session.query(
            Event.id.label('event_id'),
            Event.name.label('event_name'),
            Event.start_date.label('event_start_date'),
            Event.end_date.label('event_end_date'),
            func.sum(Activity.duration_hours).label('total_hours'),
            func.count(Activity.id).label('activities_count')
        ).join(
            Activity, Activity.event_id == Event.id
        ).join(
            Registration, Registration.activity_id == Activity.id
        ).filter(
            Registration.student_id == student_id,
            Registration.status == 'Asistió'
        ).group_by(
            Event.id, Event.name, Event.start_date, Event.end_date
        ).order_by(
            Event.start_date.desc()
        ).all()
        events_hours = []
        app_tz = current_app.config.get('APP_TIMEZONE', 'America/Mexico_City')
        for row in results:
            total_hours = float(row.total_hours or 0)
            has_credit = total_hours >= 10.0
            try:
                es = localize_naive_datetime(row.event_start_date, app_tz) if getattr(
                    row, 'event_start_date', None) is not None else None
            except Exception:
                es = None
            try:
                ee = localize_naive_datetime(row.event_end_date, app_tz) if getattr(
                    row, 'event_end_date', None) is not None else None
            except Exception:
                ee = None

            events_hours.append({
                'event_id': row.event_id,
                'event_name': row.event_name,
                'event_start_date': safe_iso(es) if es else None,
                'event_end_date': safe_iso(ee) if ee else None,
                'total_hours': total_hours,
                'activities_count': row.activities_count,
                'has_complementary_credit': has_credit
            })
        # Si no se encontraron resultados por Registration (p. ej. se registró
        # asistencia directamente en la tabla attendances), hacer un fallback
        # que calcule horas sumando las actividades relacionadas a partir de
        # la tabla Attendance.
        if len(events_hours) == 0:
            from app.models.attendance import Attendance

            attendance_results = db.session.query(
                Event.id.label('event_id'),
                Event.name.label('event_name'),
                Event.start_date.label('event_start_date'),
                Event.end_date.label('event_end_date'),
                func.sum(Activity.duration_hours).label('total_hours'),
                func.count(Activity.id).label('activities_count')
            ).join(
                Activity, Activity.event_id == Event.id
            ).join(
                Attendance, Attendance.activity_id == Activity.id
            ).filter(
                Attendance.student_id == student_id,
                Attendance.status == 'Asistió'
            ).group_by(
                Event.id, Event.name, Event.start_date, Event.end_date
            ).order_by(
                Event.start_date.desc()
            ).all()

            for row in attendance_results:
                total_hours = float(row.total_hours or 0)
                has_credit = total_hours >= 10.0
                try:
                    es = localize_naive_datetime(row.event_start_date, app_tz) if getattr(
                        row, 'event_start_date', None) is not None else None
                except Exception:
                    es = None
                try:
                    ee = localize_naive_datetime(row.event_end_date, app_tz) if getattr(
                        row, 'event_end_date', None) is not None else None
                except Exception:
                    ee = None

                events_hours.append({
                    'event_id': row.event_id,
                    'event_name': row.event_name,
                    'event_start_date': safe_iso(es) if es else None,
                    'event_end_date': safe_iso(ee) if ee else None,
                    'total_hours': total_hours,
                    'activities_count': row.activities_count,
                    'has_complementary_credit': has_credit
                })

        return jsonify({
            'student': student_schema.dump(student),
            'events_hours': events_hours
        }), 200

    except Exception as e:
        return jsonify({'message': 'Error al calcular horas por evento', 'error': str(e)}), 500


# Obtener detalle de participación de un estudiante en un evento específico
@students_bp.route('/<int:student_id>/event/<int:event_id>/details', methods=['GET'])
def get_student_event_details(student_id, event_id):
    """
    Obtiene el detalle cronológico de participación del estudiante en un evento.
    Incluye todas las actividades registradas y su status.
    """
    try:
        student = db.session.get(Student, student_id)
        if not student:
            return jsonify({'message': 'Estudiante no encontrado'}), 404

        from app.models.event import Event
        event = db.session.get(Event, event_id)
        if not event:
            return jsonify({'message': 'Evento no encontrado'}), 404

        from app.models.registration import Registration
        from app.models.activity import Activity

        # Obtener todas las registraciones del estudiante para este evento
        registrations = db.session.query(Registration).join(
            Activity, Activity.id == Registration.activity_id
        ).filter(
            Registration.student_id == student_id,
            Activity.event_id == event_id
        ).order_by(
            Activity.start_datetime.asc()
        ).all()

        activities_detail = []
        total_confirmed_hours = 0.0

        for reg in registrations:
            # Avoid direct attribute access that some static analyzers flag.
            # Prefer safe resolution via relationship if present, otherwise load by FK.
            try:
                activity = getattr(reg, 'activity', None) or db.session.get(
                    Activity, getattr(reg, 'activity_id', None))
            except Exception:
                activity = None

            # If activity could not be resolved for any reason, skip this registration
            # to avoid attribute access on None.
            if not activity:
                continue
            hours = float(activity.duration_hours or 0)

            if reg.status == 'Asistió':
                total_confirmed_hours += hours

            try:
                sdt = localize_naive_datetime(activity.start_datetime, current_app.config.get(
                    'APP_TIMEZONE', 'America/Mexico_City')) if getattr(activity, 'start_datetime', None) is not None else None
            except Exception:
                sdt = None
            try:
                edt = localize_naive_datetime(activity.end_datetime, current_app.config.get(
                    'APP_TIMEZONE', 'America/Mexico_City')) if getattr(activity, 'end_datetime', None) is not None else None
            except Exception:
                edt = None

            # normalize reg dates
            try:
                reg_dt = localize_naive_datetime(reg.registration_date, current_app.config.get(
                    'APP_TIMEZONE', 'America/Mexico_City')) if getattr(reg, 'registration_date', None) else None
            except Exception:
                reg_dt = None
            try:
                conf_dt = localize_naive_datetime(reg.confirmation_date, current_app.config.get(
                    'APP_TIMEZONE', 'America/Mexico_City')) if getattr(reg, 'confirmation_date', None) else None
            except Exception:
                conf_dt = None

            activities_detail.append({
                'registration_id': reg.id,
                'activity_id': activity.id,
                'activity_name': activity.name,
                'activity_type': activity.activity_type,
                'start_datetime': safe_iso(sdt) if sdt else None,
                'end_datetime': safe_iso(edt) if edt else None,
                'duration_hours': hours,
                'location': activity.location,
                'status': reg.status,
                'registration_date': safe_iso(reg_dt) if reg_dt else None,
                'confirmation_date': safe_iso(conf_dt) if conf_dt else None,
            })

        has_credit = total_confirmed_hours >= 10.0

        # ---- Integrar registros desde Attendance (walk-ins o asistencias directas) ----
        try:
            from app.models.attendance import Attendance

            # Mapear activities_detail por activity_id para facilitar actualizaciones
            activity_index = {a['activity_id']: idx for idx,
                              a in enumerate(activities_detail)}

            attendance_rows = db.session.query(Attendance).join(
                Activity, Activity.id == Attendance.activity_id
            ).filter(
                Attendance.student_id == student_id,
                Activity.event_id == event_id
            ).all()

            for att in attendance_rows:
                # Resolver la actividad
                try:
                    activity = getattr(att, 'activity', None) or db.session.get(
                        Activity, getattr(att, 'activity_id', None))
                except Exception:
                    activity = None

                if not activity:
                    continue

                hours = float(activity.duration_hours or 0)

                # Si ya existe una entrada por registration, actualizar estado/horas
                if activity.id in activity_index:
                    idx = activity_index[activity.id]
                    existing = activities_detail[idx]
                    # Si la asistencia confirma la participación y el registro no lo hacía,
                    # actualizar el estado y sumar las horas al total confirmado.
                    if att.status == 'Asistió' and existing.get('status') != 'Asistió':
                        existing['status'] = 'Asistió'
                        total_confirmed_hours += hours
                    # Añadir metadatos de attendance si procede
                    existing['attendance_id'] = att.id
                    existing['attendance_percentage'] = getattr(
                        att, 'attendance_percentage', None)
                    existing['check_in_time'] = safe_iso(att.check_in_time) if getattr(
                        att, 'check_in_time', None) else None
                    existing['check_out_time'] = safe_iso(att.check_out_time) if getattr(
                        att, 'check_out_time', None) else None
                else:
                    # Entrada basada únicamente en Attendance
                    try:
                        sdt = localize_naive_datetime(activity.start_datetime, current_app.config.get(
                            'APP_TIMEZONE', 'America/Mexico_City')) if getattr(activity, 'start_datetime', None) is not None else None
                    except Exception:
                        sdt = None
                    try:
                        edt = localize_naive_datetime(activity.end_datetime, current_app.config.get(
                            'APP_TIMEZONE', 'America/Mexico_City')) if getattr(activity, 'end_datetime', None) is not None else None
                    except Exception:
                        edt = None

                    att_entry = {
                        'registration_id': att.id,
                        'activity_id': activity.id,
                        'activity_name': activity.name,
                        'activity_type': activity.activity_type,
                        'start_datetime': safe_iso(sdt) if sdt else None,
                        'end_datetime': safe_iso(edt) if edt else None,
                        'duration_hours': hours,
                        'location': activity.location,
                        'status': att.status,
                        'registration_date': None,
                        'confirmation_date': None,
                        'attendance_id': att.id,
                        'attendance_percentage': getattr(att, 'attendance_percentage', None),
                        'check_in_time': safe_iso(att.check_in_time) if getattr(att, 'check_in_time', None) else None,
                        'check_out_time': safe_iso(att.check_out_time) if getattr(att, 'check_out_time', None) else None,
                    }
                    activities_detail.append(att_entry)
                    if att.status == 'Asistió':
                        total_confirmed_hours += hours

            # Reordenar activities_detail por start_datetime asc
            try:
                activities_detail.sort(
                    key=lambda x: x.get('start_datetime') or '')
            except Exception:
                pass
        except Exception:
            # No bloquear en caso de error de fallback
            pass

        try:
            ev_s = localize_naive_datetime(event.start_date, current_app.config.get(
                'APP_TIMEZONE', 'America/Mexico_City')) if getattr(event, 'start_date', None) is not None else None
        except Exception:
            ev_s = None
        try:
            ev_e = localize_naive_datetime(event.end_date, current_app.config.get(
                'APP_TIMEZONE', 'America/Mexico_City')) if getattr(event, 'end_date', None) is not None else None
        except Exception:
            ev_e = None

        return jsonify({
            'student': student_schema.dump(student),
            'event': {
                'id': event.id,
                'name': event.name,
                'start_date': safe_iso(ev_s) if ev_s else None,
                'end_date': safe_iso(ev_e) if ev_e else None,
            },
            'total_confirmed_hours': total_confirmed_hours,
            'has_complementary_credit': has_credit,
            'activities': activities_detail
        }), 200

    except Exception as e:
        return jsonify({'message': 'Error al obtener detalle del evento', 'error': str(e)}), 500


# Obtener estudiantes con 10+ horas filtrados por evento y carrera
@students_bp.route('/complementary-credits', methods=['GET'])
@jwt_required()
@require_admin
def get_students_with_complementary_credits():
    """
    Obtiene estudiantes que han acumulado 10+ horas en un evento específico,
    opcionalmente filtrados por carrera.
    """
    try:
        event_id = request.args.get('event_id', type=int)
        career = request.args.get('career', '')

        if not event_id:
            return jsonify({'message': 'event_id es requerido'}), 400

        from app.models.registration import Registration
        from app.models.activity import Activity
        from app.models.event import Event
        from sqlalchemy import func

        # Query base: estudiantes con sus horas por evento
        query = db.session.query(
            Student.id,
            Student.control_number,
            Student.full_name,
            Student.career,
            Student.email,
            func.sum(Activity.duration_hours).label('total_hours'),
            func.count(Activity.id).label('activities_count')
        ).join(
            Registration, Registration.student_id == Student.id
        ).join(
            Activity, Activity.id == Registration.activity_id
        ).filter(
            Activity.event_id == event_id,
            Registration.status == 'Asistió'
        )

        # Filtro opcional por carrera
        if career:
            query = query.filter(Student.career.ilike(f'%{career}%'))

        # Agrupar por estudiante y filtrar por horas >= 10
        query = query.group_by(
            Student.id, Student.control_number, Student.full_name,
            Student.career, Student.email
        ).having(
            func.sum(Activity.duration_hours) >= 10.0
        ).order_by(
            Student.full_name
        )

        results = query.all()

        # Obtener información del evento
        event = db.session.get(Event, event_id)
        if not event:
            return jsonify({'message': 'Evento no encontrado'}), 404

        students_list = []
        for row in results:
            students_list.append({
                'id': row.id,
                'control_number': row.control_number,
                'full_name': row.full_name,
                'career': row.career or 'Sin carrera',
                'email': row.email or 'Sin email',
                'total_hours': float(row.total_hours or 0),
                'activities_count': row.activities_count,
                'has_complementary_credit': True  # Ya filtrados por >= 10 horas
            })

        # Localizar fechas del evento de forma consistente antes de serializar
        try:
            app_tz = current_app.config.get(
                'APP_TIMEZONE', 'America/Mexico_City')
            ev_s = localize_naive_datetime(event.start_date, app_tz) if getattr(
                event, 'start_date', None) is not None else None
        except Exception:
            ev_s = None
        try:
            ev_e = localize_naive_datetime(event.end_date, app_tz) if getattr(
                event, 'end_date', None) is not None else None
        except Exception:
            ev_e = None

        return jsonify({
            'event': {
                'id': event.id,
                'name': event.name,
                'start_date': safe_iso(ev_s) if ev_s else None,
                'end_date': safe_iso(ev_e) if ev_e else None,
            },
            'students': students_list,
            'total_students': len(students_list)
        }), 200

    except Exception as e:
        return jsonify({'message': 'Error al obtener estudiantes con créditos', 'error': str(e)}), 500


# Exportar estudiantes con crédito complementario a Excel
@students_bp.route('/complementary-credits/export', methods=['GET'])
@jwt_required()
@require_admin
def export_complementary_credits():
    """
    Exporta a Excel la lista de estudiantes con 10+ horas en un evento.
    """
    try:
        event_id = request.args.get('event_id', type=int)
        career = request.args.get('career', '')

        if not event_id:
            return jsonify({'message': 'event_id es requerido'}), 400

        from app.models.registration import Registration
        from app.models.activity import Activity
        from app.models.event import Event
        from sqlalchemy import func

        # Obtener datos (misma lógica que el endpoint anterior)
        query = db.session.query(
            Student.id,
            Student.control_number,
            Student.full_name,
            Student.career,
            Student.email,
            func.sum(Activity.duration_hours).label('total_hours'),
            func.count(Activity.id).label('activities_count')
        ).join(
            Registration, Registration.student_id == Student.id
        ).join(
            Activity, Activity.id == Registration.activity_id
        ).filter(
            Activity.event_id == event_id,
            Registration.status == 'Asistió'
        )

        if career:
            query = query.filter(Student.career.ilike(f'%{career}%'))

        query = query.group_by(
            Student.id, Student.control_number, Student.full_name,
            Student.career, Student.email
        ).having(
            func.sum(Activity.duration_hours) >= 10.0
        ).order_by(
            Student.full_name
        )

        results = query.all()

        # Obtener información del evento
        event = db.session.get(Event, event_id)
        if not event:
            return jsonify({'message': 'Evento no encontrado'}), 404

        # Crear archivo Excel
        wb: Workbook = Workbook()
        ws: Any = wb.active
        ws.title = "Créditos Complementarios"

        # Estilos
        header_fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Título
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"Estudiantes con Crédito Complementario - {event.name}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center")

        # Información adicional
        ws.merge_cells('A2:G2')
        info_cell = ws['A2']
        info_cell.value = f"Generado el: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}"
        info_cell.alignment = Alignment(horizontal="center")

        if career:
            ws.merge_cells('A3:G3')
            career_cell = ws['A3']
            career_cell.value = f"Filtrado por carrera: {career}"
            career_cell.alignment = Alignment(horizontal="center")
            header_row = 5
        else:
            header_row = 4

        # Encabezados
        headers = ['No.', 'Número de Control', 'Nombre Completo',
                   'Carrera', 'Email', 'Horas Confirmadas', 'Actividades']
        for col_num, header in enumerate(headers, 1):
            cell: Any = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Datos
        for idx, row in enumerate(results, 1):
            data_row = header_row + idx
            ws.cell(row=data_row, column=1, value=idx)
            ws.cell(row=data_row, column=2, value=row.control_number)
            ws.cell(row=data_row, column=3, value=row.full_name)
            ws.cell(row=data_row, column=4, value=row.career or 'Sin carrera')
            ws.cell(row=data_row, column=5, value=row.email or 'Sin email')
            ws.cell(row=data_row, column=6, value=float(row.total_hours or 0))
            ws.cell(row=data_row, column=7, value=row.activities_count)

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15

        # Resumen al final
        summary_row = header_row + len(results) + 2
        ws.merge_cells(f'A{summary_row}:E{summary_row}')
        summary_cell: Any = ws.cell(row=summary_row, column=1)
        summary_cell.value = f"Total de estudiantes: {len(results)}"
        summary_cell.font = Font(bold=True)
        summary_cell.alignment = Alignment(horizontal="right")

        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generar nombre de archivo
        filename = f"creditos_complementarios_{event.name.replace(' ', '_')}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Localizar fechas del evento antes de devolver metadatos en la exportación
        try:
            app_tz = current_app.config.get(
                'APP_TIMEZONE', 'America/Mexico_City')
            ev_s = localize_naive_datetime(event.start_date, app_tz) if getattr(
                event, 'start_date', None) is not None else None
        except Exception:
            ev_s = None
        try:
            ev_e = localize_naive_datetime(event.end_date, app_tz) if getattr(
                event, 'end_date', None) is not None else None
        except Exception:
            ev_e = None

        # (filename ya fue generado arriba con UTC now)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'message': 'Error al exportar datos', 'error': str(e)}), 500
